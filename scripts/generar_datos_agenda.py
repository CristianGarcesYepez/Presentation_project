# crear_excel_ejemplo.py - Script para generar el archivo Excel con datos de ejemplo
import pandas as pd
from datetime import datetime, timedelta

def crear_excel_agenda_visitas():
    """Crea un archivo Excel con datos de ejemplo para la agenda de visitas"""
    
    # Obtener fecha actual y calcular fechas para las próximas 4 semanas
    hoy = datetime.now()
    
    # Datos de ejemplo distribuidos en las próximas 4 semanas
    datos_ejemplo = [
        # Semana 1
        {'Fecha': '2024-06-03', 'Tarea': 'California', 'Duración': '4h', 'Descripción': 'Introducción Juvenil', 'Estado': 'Programado', 'Zona': 'Norte', 'Responsable': 'Juan Pérez'},
        {'Fecha': '2024-06-04', 'Tarea': 'California', 'Duración': '4h', 'Descripción': 'Enganche', 'Estado': 'Programado', 'Zona': 'Norte', 'Responsable': 'Juan Pérez'},
        {'Fecha': '2024-06-05', 'Tarea': 'California', 'Duración': '4h', 'Descripción': 'Seguimiento', 'Estado': 'Completado', 'Zona': 'Norte', 'Responsable': 'Juan Pérez'},
        
        # Semana 2
        {'Fecha': '2024-06-10', 'Tarea': 'Taura C', 'Duración': '6h', 'Descripción': 'Capacitación Insigne', 'Estado': 'Programado', 'Zona': 'Centro', 'Responsable': 'María García'},
        {'Fecha': '2024-06-11', 'Tarea': 'Churute', 'Duración': '5h', 'Descripción': 'Muestreo de Pesos', 'Estado': 'En Progreso', 'Zona': 'Sur', 'Responsable': 'Carlos López'},
        {'Fecha': '2024-06-12', 'Tarea': 'Taura', 'Duración': '3h', 'Descripción': 'Capacitación INSIGNE', 'Estado': 'Programado', 'Zona': 'Centro', 'Responsable': 'María García'},
        {'Fecha': '2024-06-13', 'Tarea': 'Taura', 'Duración': '3h', 'Descripción': 'Evaluación Final', 'Estado': 'Programado', 'Zona': 'Centro', 'Responsable': 'María García'},
        
        # Semana 3
        {'Fecha': '2024-06-17', 'Tarea': 'Durán', 'Duración': '4h', 'Descripción': 'Capacitación General', 'Estado': 'Programado', 'Zona': 'Este', 'Responsable': 'Ana Rodríguez'},
        {'Fecha': '2024-06-18', 'Tarea': 'Samborondón', 'Duración': '3h', 'Descripción': 'Seguimiento Proyecto', 'Estado': 'Programado', 'Zona': 'Norte', 'Responsable': 'Luis Morales'},
        {'Fecha': '2024-06-19', 'Tarea': 'Milagro', 'Duración': '5h', 'Descripción': 'Evaluación Mensual', 'Estado': 'Programado', 'Zona': 'Interior', 'Responsable': 'Patricia Silva'},
        {'Fecha': '2024-06-20', 'Tarea': 'Yaguachi', 'Duración': '4h', 'Descripción': 'Capacitación Técnica', 'Estado': 'Programado', 'Zona': 'Interior', 'Responsable': 'Roberto Castro'},
        
        # Semana 4
        {'Fecha': '2024-06-24', 'Tarea': 'Guayaquil Centro', 'Duración': '2h', 'Descripción': 'Reunión Coordinación', 'Estado': 'Programado', 'Zona': 'Centro', 'Responsable': 'Juan Pérez'},
        {'Fecha': '2024-06-25', 'Tarea': 'Pascuales', 'Duración': '3h', 'Descripción': 'Planificación Semanal', 'Estado': 'Programado', 'Zona': 'Oeste', 'Responsable': 'Carmen Vega'},
        {'Fecha': '2024-06-26', 'Tarea': 'La Libertad', 'Duración': '4h', 'Descripción': 'Revisión Resultados', 'Estado': 'Programado', 'Zona': 'Costa', 'Responsable': 'Diego Herrera'},
        {'Fecha': '2024-06-27', 'Tarea': 'Salinas', 'Duración': '5h', 'Descripción': 'Capacitación Avanzada', 'Estado': 'Programado', 'Zona': 'Costa', 'Responsable': 'Diego Herrera'},
        {'Fecha': '2024-06-28', 'Tarea': 'Playas', 'Duración': '3h', 'Descripción': 'Seguimiento Costa', 'Estado': 'Programado', 'Zona': 'Costa', 'Responsable': 'Diego Herrera'},
        
        # Datos adicionales para mostrar variedad
        {'Fecha': '2024-07-01', 'Tarea': 'El Triunfo', 'Duración': '4h', 'Descripción': 'Evaluación Trimestral', 'Estado': 'Programado', 'Zona': 'Interior', 'Responsable': 'Patricia Silva'},
        {'Fecha': '2024-07-02', 'Tarea': 'Naranjal', 'Duración': '3h', 'Descripción': 'Capacitación Básica', 'Estado': 'Programado', 'Zona': 'Sur', 'Responsable': 'Carlos López'},
        {'Fecha': '2024-07-03', 'Tarea': 'Balao', 'Duración': '5h', 'Descripción': 'Proyecto Especial', 'Estado': 'Programado', 'Zona': 'Sur', 'Responsable': 'Carlos López'},
        {'Fecha': '2024-07-04', 'Tarea': 'Pedro Carbo', 'Duración': '4h', 'Descripción': 'Implementación Nueva Metodología', 'Estado': 'Programado', 'Zona': 'Interior', 'Responsable': 'Roberto Castro'},
        {'Fecha': '2024-07-05', 'Tarea': 'Colimes', 'Duración': '3h', 'Descripción': 'Seguimiento Mensual', 'Estado': 'Programado', 'Zona': 'Interior', 'Responsable': 'Roberto Castro'},
    ]
    
    # Crear DataFrame
    df = pd.DataFrame(datos_ejemplo)
    
    # Convertir la columna Fecha a datetime
    df['Fecha'] = pd.to_datetime(df['Fecha'])
    
    # Ordenar por fecha
    df = df.sort_values('Fecha').reset_index(drop=True)
    
    # Guardar en Excel con formato
    with pd.ExcelWriter('agenda_visitas_data.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Agenda_Visitas', index=False)
        
        # Obtener el workbook y worksheet para formatear
        workbook = writer.book
        worksheet = writer.sheets['Agenda_Visitas']
        
        # Ajustar ancho de columnas
        column_widths = {
            'A': 12,  # Fecha
            'B': 20,  # Tarea
            'C': 10,  # Duración
            'D': 35,  # Descripción
            'E': 15,  # Estado
            'F': 15,  # Zona
            'G': 20   # Responsable
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # Formatear encabezados
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Aplicar formato a encabezados
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Formatear fechas
        from openpyxl.styles import NamedStyle
        date_style = NamedStyle(name='date_style', number_format='DD/MM/YYYY')
        
        for row in range(2, len(df) + 2):
            worksheet[f'A{row}'].style = date_style
            worksheet[f'A{row}'].alignment = Alignment(horizontal="center")
        
        # Formatear estados con colores
        for row in range(2, len(df) + 2):
            estado_cell = worksheet[f'E{row}']
            if estado_cell.value == 'Completado':
                estado_cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
                estado_cell.font = Font(color="FFFFFF", bold=True)
            elif estado_cell.value == 'En Progreso':
                estado_cell.fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
                estado_cell.font = Font(color="FFFFFF", bold=True)
            elif estado_cell.value == 'Programado':
                estado_cell.fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                estado_cell.font = Font(color="FFFFFF", bold=True)
    
    print(f"✅ Archivo 'agenda_visitas_data.xlsx' creado exitosamente!")
    print(f"📊 {len(df)} tareas agregadas")
    print(f"📅 Fechas desde {df['Fecha'].min().strftime('%d/%m/%Y')} hasta {df['Fecha'].max().strftime('%d/%m/%Y')}")
    print(f"🎯 {df['Estado'].value_counts().to_dict()}")
    
    return df

if __name__ == "__main__":
    crear_excel_agenda_visitas()